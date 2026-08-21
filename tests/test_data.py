import tempfile
import unittest
from pathlib import Path

import openpyxl

from src.data import (
    VLESample,
    build_split_plan,
    classify_quality_flags,
    deduplicate_samples,
    grouped_holdout_and_folds,
    infer_experiment_modes,
    load_vle_dataset,
    pure_anchor_temperatures,
    retain_pure_anchored_systems,
)


def sample(smiles: tuple[str, ...], index: int) -> VLESample:
    n = len(smiles)
    return VLESample(
        smiles=smiles,
        names=smiles,
        temperature_k=330.0 + index,
        pressure_kpa=101.325,
        liquid_composition=tuple([1.0 / n] * n),
        vapor_composition=tuple([1.0 / n] * n),
        quality_weight=1.0,
        quality_status="passed",
        source="synthetic.xlsx",
        doi=f"doi:{index}",
    )


class GroupedSplitTests(unittest.TestCase):
    def test_experiment_mode_is_inferred_from_repeated_series_conditions(self) -> None:
        rows = [
            VLESample(
                smiles=("A", "B"),
                names=("A", "B"),
                temperature_k=temperature,
                pressure_kpa=pressure,
                liquid_composition=(0.5, 0.5),
                vapor_composition=(0.5, 0.5),
                quality_weight=1.0,
                quality_status="passed",
                source="series.xlsx",
                doi=doi,
            )
            for doi, temperatures, pressures in (
                ("isobaric", (320.0, 330.0, 340.0), (101.325, 101.325, 101.325)),
                ("isothermal", (350.0, 350.0, 350.0), (80.0, 90.0, 100.0)),
                ("full", (360.0, 370.0), (110.0, 120.0)),
            )
            for temperature, pressure in zip(temperatures, pressures)
        ]

        inferred = infer_experiment_modes(rows, minimum_series_points=3)

        modes_by_doi = {
            doi: {row.experiment_mode for row in inferred if row.doi == doi}
            for doi in ("isobaric", "isothermal", "full")
        }
        self.assertEqual(modes_by_doi["isobaric"], {"isobaric"})
        self.assertEqual(modes_by_doi["isothermal"], {"isothermal"})
        self.assertEqual(modes_by_doi["full"], {"full_state"})
        inferred_rows = [row for row in inferred if row.doi != "full"]
        self.assertTrue(all(row.experiment_mode_confidence > 0.5 for row in inferred_rows))

    def test_explicit_mode_wins_and_missing_doi_falls_back_to_full_state(self) -> None:
        explicit = VLESample(
            smiles=("A", "B"),
            names=("A", "B"),
            temperature_k=350.0,
            pressure_kpa=80.0,
            liquid_composition=(0.5, 0.5),
            vapor_composition=(0.5, 0.5),
            quality_weight=1.0,
            quality_status="passed",
            source="series.xlsx",
            doi="explicit",
            experiment_mode="isothermal",
            experiment_mode_confidence=1.0,
        )
        missing_doi = [
            VLESample(
                smiles=("C", "D"),
                names=("C", "D"),
                temperature_k=temperature,
                pressure_kpa=101.325,
                liquid_composition=(0.5, 0.5),
                vapor_composition=(0.5, 0.5),
                quality_weight=1.0,
                quality_status="passed",
                source="series.xlsx",
                doi="",
            )
            for temperature in (320.0, 330.0, 340.0)
        ]

        inferred = infer_experiment_modes([explicit, *missing_doi])

        self.assertEqual(inferred[0].experiment_mode, "isothermal")
        self.assertEqual(inferred[0].experiment_mode_confidence, 1.0)
        self.assertTrue(all(row.experiment_mode == "full_state" for row in inferred[1:]))

    def test_low_confidence_inference_falls_back_to_full_state(self) -> None:
        rows = [
            VLESample(
                smiles=("A", "B"),
                names=("A", "B"),
                temperature_k=temperature,
                pressure_kpa=pressure,
                liquid_composition=(0.5, 0.5),
                vapor_composition=(0.5, 0.5),
                quality_weight=1.0,
                quality_status="passed",
                source="ambiguous.xlsx",
                doi="ambiguous",
            )
            for temperature, pressure in (
                (320.0, 100.0),
                (320.0, 100.0),
                (330.0, 100.0),
            )
        ]

        inferred = infer_experiment_modes(rows)

        self.assertEqual(inferred[0].experiment_mode, "full_state")
        self.assertAlmostEqual(inferred[0].experiment_mode_confidence, 0.6)

    def test_dataset_load_reports_rejection_and_deduplication_counts(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "name1", "formula1", "smiles1", "name2", "formula2", "smiles2",
                    "check1", "check2", "P", "T", "x1", "y1", "doi",
                ]
            )
            valid = ["A", "", "A", "B", "", "B", 1, 1, 760.0, 30.0, 0.4, 0.5, "d"]
            sheet.append(valid)
            sheet.append(valid)
            sheet.append(["A", "", "", "B", "", "B", 1, 1, 760.0, 30.0, 0.4, 0.5, "d"])
            sheet.append(["A", "", "A", "B", "", "B", 0, -1, 760.0, 30.0, 0.4, 0.5, "d"])
            sheet.append(["A", "", "A", "B", "", "B", 1, 1, 760.0, 30.0, 1.5, 0.5, "d"])
            sheet.append(["A", "", "A", "B", "", "B", 1, 1, 20000.0, 30.0, 0.4, 0.5, "d"])
            workbook.save(root / "binary.xlsx")

            result = load_vle_dataset(root, failed_weight=0.0, max_pressure_kpa=500.0)

        self.assertEqual(len(result.samples), 1)
        self.assertEqual(result.audit.raw_rows, 6)
        self.assertEqual(result.audit.rejected["missing_smiles"], 1)
        self.assertEqual(result.audit.rejected["failed_quality"], 1)
        self.assertEqual(result.audit.rejected["invalid_composition"], 1)
        self.assertEqual(result.audit.rejected["pressure_above_limit"], 1)
        self.assertEqual(result.audit.duplicates_removed, 1)

    def test_dataset_loader_prefers_explicit_mode_column(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "name1", "formula1", "smiles1", "name2", "formula2", "smiles2",
                    "check1", "check2", "P", "T", "x1", "y1", "doi",
                    "experiment_mode",
                ]
            )
            for temperature in (30.0, 40.0, 50.0):
                sheet.append(
                    [
                        "A", "", "A", "B", "", "B", 1, 1, 760.0,
                        temperature, 0.4, 0.5, "explicit-series", "isothermal",
                    ]
                )
            workbook.save(root / "binary.xlsx")

            result = load_vle_dataset(root)

        self.assertEqual({row.experiment_mode for row in result.samples}, {"isothermal"})
        self.assertEqual(
            {row.experiment_mode_confidence for row in result.samples},
            {1.0},
        )

    def test_explicit_diagram_aliases_have_correct_thermodynamic_direction(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(
                [
                    "name1", "formula1", "smiles1", "name2", "formula2", "smiles2",
                    "check1", "check2", "P", "T", "x1", "y1", "doi",
                    "experiment_mode",
                ]
            )
            sheet.append(
                ["A", "", "A", "B", "", "B", 1, 1, 760.0, 30.0, 0.4, 0.5, "txy", "T-x-y"]
            )
            sheet.append(
                ["A", "", "A", "B", "", "B", 1, 1, 700.0, 40.0, 0.4, 0.5, "pxy", "P-x-y"]
            )
            sheet.append(
                ["A", "", "A", "B", "", "B", 1, 1, 650.0, 50.0, 0.4, 0.5, "bad", "isothremal"]
            )
            workbook.save(root / "binary.xlsx")

            result = load_vle_dataset(root)

        self.assertEqual(
            {row.doi: row.experiment_mode for row in result.samples},
            {"txy": "isobaric", "pxy": "isothermal"},
        )
        self.assertEqual(result.audit.rejected["invalid_experiment_mode"], 1)

    def test_quality_codes_match_workbook_convention(self) -> None:
        self.assertEqual(classify_quality_flags((1, -1), 0.1), ("passed", 1.0))
        self.assertEqual(classify_quality_flags((-1, -1), 0.1), ("unverified", 0.5))
        self.assertEqual(classify_quality_flags((0, -1), 0.1), ("failed", 0.1))
        self.assertEqual(classify_quality_flags((1, 0), 0.1), ("failed", 0.1))

    def test_reversed_component_duplicate_is_removed(self) -> None:
        original = sample(("A", "B"), 1)
        reversed_duplicate = VLESample(
            smiles=("B", "A"),
            names=("B", "A"),
            temperature_k=original.temperature_k,
            pressure_kpa=original.pressure_kpa,
            liquid_composition=tuple(reversed(original.liquid_composition)),
            vapor_composition=tuple(reversed(original.vapor_composition)),
            quality_weight=original.quality_weight,
            quality_status=original.quality_status,
            source="aggregate.xlsx",
            doi=original.doi,
        )

        unique = deduplicate_samples([original, reversed_duplicate])

        self.assertEqual(unique, [original])

    def test_unanchored_pure_property_system_is_removed(self) -> None:
        anchored_rows = [
            VLESample(
                smiles=("A", "B"),
                names=("A", "B"),
                temperature_k=temperature,
                pressure_kpa=100.0,
                liquid_composition=composition,
                vapor_composition=composition,
                quality_weight=1.0,
                quality_status="passed",
                source="synthetic.xlsx",
                doi="synthetic",
            )
            for temperature in (300.0, 320.0)
            for composition in ((1.0, 0.0), (0.0, 1.0))
        ]
        unanchored = sample(("A", "C"), 9)

        retained = retain_pure_anchored_systems(
            anchored_rows + [unanchored], minimum_temperatures=2
        )

        self.assertEqual(retained, anchored_rows)

    def test_systems_never_cross_holdout_or_fold_seams(self) -> None:
        systems = [
            ("A", "B"),
            ("A", "C"),
            ("B", "C"),
            ("A", "D"),
            ("B", "D"),
            ("C", "D"),
            ("A", "B", "C"),
            ("A", "B", "D"),
            ("A", "C", "D"),
            ("B", "C", "D"),
            ("A", "B", "E"),
            ("A", "C", "E"),
        ]
        samples = [
            sample(system if i % 2 == 0 else tuple(reversed(system)), i)
            for i, system in enumerate(systems)
            for _ in range(3)
        ]

        plan = grouped_holdout_and_folds(samples, test_fraction=0.2, folds=3, seed=7)

        test_systems = {row.system_key for row in plan.test}
        cv_systems = {row.system_key for row in plan.cv}
        self.assertTrue(test_systems.isdisjoint(cv_systems))
        self.assertEqual({row.component_count for row in plan.test}, {2, 3})
        for fold in plan.folds:
            train_systems = {row.system_key for row in fold.train}
            validation_systems = {row.system_key for row in fold.validation}
            self.assertTrue(train_systems.isdisjoint(validation_systems))
            self.assertEqual(train_systems | validation_systems, cv_systems)
            self.assertEqual({row.component_count for row in fold.validation}, {2, 3})

    def test_every_training_partition_keeps_pure_property_anchors(self) -> None:
        anchor_systems = [("A", "B"), ("C", "D"), ("A", "E")]
        anchors = [
            VLESample(
                smiles=system,
                names=system,
                temperature_k=temperature,
                pressure_kpa=100.0,
                liquid_composition=composition,
                vapor_composition=composition,
                quality_weight=1.0,
                quality_status="passed",
                source="anchors.xlsx",
                doi="pure",
            )
            for system in anchor_systems
            for temperature in (300.0, 320.0)
            for composition in ((1.0, 0.0), (0.0, 1.0))
        ]
        evaluation_systems = [
            ("A", "C"),
            ("A", "D"),
            ("B", "C"),
            ("B", "D"),
            ("B", "E"),
            ("C", "E"),
            ("A", "B", "C"),
            ("A", "C", "D"),
            ("B", "D", "E"),
            ("A", "D", "E"),
        ]
        rows = anchors + [sample(system, index + 100) for index, system in enumerate(evaluation_systems)]

        plan = grouped_holdout_and_folds(
            rows,
            test_fraction=0.2,
            folds=2,
            seed=3,
            minimum_anchor_temperatures=2,
        )

        all_molecules = {smile for row in rows for smile in row.smiles}
        for training_rows in [plan.cv, *(fold.train for fold in plan.folds)]:
            temperatures = pure_anchor_temperatures(training_rows)
            self.assertTrue(all(len(temperatures.get(smile, set())) >= 2 for smile in all_molecules))
        reference = set(plan.anchor_reference_systems)
        self.assertTrue(reference)
        self.assertTrue(reference.isdisjoint({row.system_key for row in plan.test}))
        for fold in plan.folds:
            self.assertTrue(reference.isdisjoint({row.system_key for row in fold.validation}))

    def test_holdout_mode_returns_disjoint_train_validation_and_test(self) -> None:
        systems = [
            ("A", "B"), ("A", "C"), ("A", "D"), ("B", "C"),
            ("B", "D"), ("C", "D"), ("A", "E"), ("B", "E"),
            ("A", "B", "C"), ("A", "B", "D"), ("A", "C", "D"),
            ("B", "C", "D"), ("A", "B", "E"), ("A", "C", "E"),
        ]
        rows = [sample(system, index) for index, system in enumerate(systems)]

        plan = build_split_plan(
            rows,
            mode="holdout",
            test_fraction=0.2,
            validation_fraction=0.2,
            folds=5,
            seed=9,
        )

        self.assertEqual(plan.mode, "holdout")
        self.assertEqual(len(plan.folds), 1)
        split = plan.folds[0]
        train = {row.system_key for row in split.train}
        validation = {row.system_key for row in split.validation}
        test = {row.system_key for row in plan.test}
        self.assertTrue(train.isdisjoint(validation))
        self.assertTrue(train.isdisjoint(test))
        self.assertTrue(validation.isdisjoint(test))
        self.assertEqual(train | validation | test, {row.system_key for row in rows})

    def test_kfold_mode_uses_requested_number_of_folds(self) -> None:
        systems = [(f"A{i}", f"B{i}") for i in range(12)]
        rows = [sample(system, index) for index, system in enumerate(systems)]

        plan = build_split_plan(rows, mode="kfold", folds=5, seed=2)

        self.assertEqual(plan.mode, "kfold")
        self.assertEqual(len(plan.folds), 5)


if __name__ == "__main__":
    unittest.main()
