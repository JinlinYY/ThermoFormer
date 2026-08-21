"""Shared, read-only parsing and identity utilities for the VLE dataset audit."""

from __future__ import annotations

import hashlib
import math
import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from itertools import combinations
from pathlib import Path
from typing import Any, Sequence

import numpy as np
import openpyxl
import pandas as pd
from rdkit import Chem

MMHG_TO_KPA = 0.133322368
SEED = 42


@dataclass(frozen=True)
class WorkbookData:
    path: Path
    data_sheet: str
    component_count: int
    frame: pd.DataFrame
    dictionary: pd.DataFrame
    schema: dict[str, Any]


@dataclass(frozen=True)
class UnifiedData:
    workbooks: tuple[WorkbookData, ...]
    records: pd.DataFrame
    occurrences: pd.DataFrame


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_text(value: object) -> str:
    if value is None or (not isinstance(value, (list, tuple, dict)) and pd.isna(value)):
        return ""
    text = unicodedata.normalize("NFKC", str(value or "")).strip().casefold()
    return re.sub(r"\s+", " ", text)


def _canonicalize_smiles(value: object) -> tuple[str, str, str]:
    if value is None or pd.isna(value):
        return "", "", "missing"
    raw = str(value or "").strip()
    if not raw:
        return "", "", "missing"
    return _canonicalize_smiles_text(raw)


@lru_cache(maxsize=None)
def _canonicalize_smiles_text(raw: str) -> tuple[str, str, str]:
    mol = Chem.MolFromSmiles(raw)
    if mol is None:
        return "", "", "parse_failed"
    canonical = Chem.MolToSmiles(mol, canonical=True, isomericSmiles=True)
    try:
        inchikey = Chem.MolToInchiKey(mol)
    except Exception:
        inchikey = ""
    return canonical, inchikey, "parsed"


def _fallback_key(name: object, formula: object) -> str:
    return f"{normalize_text(name)}|{normalize_text(formula)}"


def _fallback_id(key: str) -> str:
    return f"fallback:{hashlib.sha1(key.encode('utf-8')).hexdigest()[:16]}"


def _sheet_headers(sheet: openpyxl.worksheet.worksheet.Worksheet) -> list[str]:
    row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(value or "").strip() for value in row]


def _workbook_schema(path: Path) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=False)
    try:
        sheets = []
        for sheet in workbook.worksheets:
            formula_count = sum(
                1
                for row in sheet.iter_rows()
                for cell in row
                if cell.data_type == "f"
            )
            hidden_rows = sum(
                bool(dimension.hidden) for dimension in sheet.row_dimensions.values()
            )
            hidden_columns = sum(
                bool(dimension.hidden) for dimension in sheet.column_dimensions.values()
            )
            sheets.append(
                {
                    "name": sheet.title,
                    "state": sheet.sheet_state,
                    "rows_including_header": sheet.max_row,
                    "columns": sheet.max_column,
                    "headers": _sheet_headers(sheet),
                    "merged_ranges": len(sheet.merged_cells.ranges),
                    "formula_cells": formula_count,
                    "hidden_rows": hidden_rows,
                    "hidden_columns": hidden_columns,
                    "data_validations": len(sheet.data_validations.dataValidation),
                    "tables": len(sheet.tables),
                }
            )
        return {
            "filename": path.name,
            "path": str(path.resolve()),
            "format": path.suffix.lower(),
            "size_bytes": path.stat().st_size,
            "sha256": sha256(path),
            "sheet_count": len(workbook.sheetnames),
            "defined_names": len(workbook.defined_names),
            "sheets": sheets,
        }
    finally:
        workbook.close()


def discover_workbooks(dataset_root: Path) -> tuple[WorkbookData, ...]:
    files = sorted(dataset_root.rglob("*.xlsx"))
    if len(files) != 2:
        raise RuntimeError(
            f"Expected exactly two .xlsx datasets under {dataset_root}; found {len(files)}"
        )
    discovered = []
    for path in files:
        excel = pd.ExcelFile(path, engine="openpyxl")
        sheet_headers: dict[str, list[str]] = {}
        for sheet_name in excel.sheet_names:
            header = pd.read_excel(path, sheet_name=sheet_name, nrows=0).columns
            sheet_headers[sheet_name] = [str(value).strip() for value in header]
        candidates = [
            name
            for name, columns in sheet_headers.items()
            if "smiles1" in columns
            and "pressure_mmhg" in columns
            and "temperature_c" in columns
            and "x1" in columns
            and "y1" in columns
        ]
        if len(candidates) != 1:
            raise RuntimeError(
                f"Could not uniquely identify the VLE data sheet in {path.name}: {candidates}"
            )
        data_sheet = candidates[0]
        frame = pd.read_excel(path, sheet_name=data_sheet, engine="openpyxl")
        component_count = sum(
            f"smiles{index}" in frame.columns for index in range(1, 5)
        )
        if component_count not in (2, 3):
            raise RuntimeError(
                f"Only binary/ternary VLE tables are supported; {path.name} has {component_count}"
            )
        dictionary = (
            pd.read_excel(path, sheet_name="data_dictionary", engine="openpyxl")
            if "data_dictionary" in excel.sheet_names
            else pd.DataFrame(columns=["field", "description", "unit_or_codes"])
        )
        schema = _workbook_schema(path)
        schema["data_sheet"] = data_sheet
        schema["component_count"] = component_count
        schema["main_rows"] = int(len(frame))
        schema["main_columns"] = int(frame.shape[1])
        schema["main_dtypes"] = {name: str(dtype) for name, dtype in frame.dtypes.items()}
        schema["main_missing"] = {
            name: int(count) for name, count in frame.isna().sum().items()
        }
        schema["main_exact_duplicate_rows"] = int(frame.duplicated().sum())
        discovered.append(
            WorkbookData(
                path=path,
                data_sheet=data_sheet,
                component_count=component_count,
                frame=frame,
                dictionary=dictionary,
                schema=schema,
            )
        )
    return tuple(sorted(discovered, key=lambda item: item.component_count))


def _quality_status(first: object, second: object) -> str:
    values = {int(value) for value in (first, second) if pd.notna(value)}
    if 0 in values:
        return "failed"
    if 1 in values:
        return "passed"
    return "not_evaluated"


def _component_occurrences(workbooks: Sequence[WorkbookData]) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    for workbook in workbooks:
        dataset = "Binary" if workbook.component_count == 2 else "Ternary"
        for index, raw in workbook.frame.iterrows():
            for position in range(1, workbook.component_count + 1):
                name = raw[f"component_{position}_original_name"]
                formula = raw[f"component_{position}_formula"]
                raw_smiles = raw[f"smiles{position}"]
                canonical, inchikey, status = _canonicalize_smiles(raw_smiles)
                rows.append(
                    {
                        "dataset": dataset,
                        "source_file": workbook.path.name,
                        "record_id": f"{dataset.lower()}:{index + 1:06d}",
                        "excel_row": int(index + 2),
                        "component_position": position,
                        "original_name": str(name or "").strip(),
                        "formula": str(formula or "").strip(),
                        "raw_smiles": "" if pd.isna(raw_smiles) else str(raw_smiles).strip(),
                        "canonical_smiles": canonical,
                        "inchikey": inchikey,
                        "smiles_status": status,
                        "fallback_key": _fallback_key(name, formula),
                    }
                )
    occurrences = pd.DataFrame(rows)
    canonical_by_fallback = (
        occurrences.loc[occurrences["canonical_smiles"] != ""]
        .groupby("fallback_key")["canonical_smiles"]
        .agg(lambda values: tuple(sorted(set(values))))
        .to_dict()
    )

    def resolve(row: pd.Series) -> tuple[str, str]:
        if row["canonical_smiles"]:
            return f"smiles:{row['canonical_smiles']}", "canonical_smiles"
        candidates = canonical_by_fallback.get(row["fallback_key"], ())
        if len(candidates) == 1:
            return f"smiles:{candidates[0]}", "name_formula_linked_to_smiles"
        return _fallback_id(row["fallback_key"]), "name_formula_fallback"

    resolved = occurrences.apply(resolve, axis=1, result_type="expand")
    resolved.columns = ["component_id", "identity_method"]
    return pd.concat([occurrences, resolved], axis=1)


def load_unified_data(dataset_root: Path) -> UnifiedData:
    workbooks = discover_workbooks(dataset_root)
    occurrences = _component_occurrences(workbooks)
    occurrence_lookup = {
        (row.dataset, row.record_id, row.component_position): row
        for row in occurrences.itertuples(index=False)
    }
    records: list[dict[str, Any]] = []
    for workbook in workbooks:
        dataset = "Binary" if workbook.component_count == 2 else "Ternary"
        for index, raw in workbook.frame.iterrows():
            record_id = f"{dataset.lower()}:{index + 1:06d}"
            component_rows = [
                occurrence_lookup[(dataset, record_id, position)]
                for position in range(1, workbook.component_count + 1)
            ]
            component_ids = [str(row.component_id) for row in component_rows]
            independent_x = [float(raw[f"x{position}"]) for position in range(1, workbook.component_count)]
            independent_y = [float(raw[f"y{position}"]) for position in range(1, workbook.component_count)]
            x = independent_x + [1.0 - sum(independent_x)]
            y = independent_y + [1.0 - sum(independent_y)]
            order = sorted(range(workbook.component_count), key=lambda i: component_ids[i])
            ordered_ids = [component_ids[i] for i in order]
            ordered_x = [x[i] for i in order]
            ordered_y = [y[i] for i in order]
            temperature_c = float(raw["temperature_c"])
            pressure_mmhg = float(raw["pressure_mmhg"])
            record: dict[str, Any] = {
                "dataset": dataset,
                "component_count": workbook.component_count,
                "source_file": workbook.path.name,
                "record_id": record_id,
                "excel_row": int(index + 2),
                "doi": str(raw.get("doi", "") or "").strip(),
                "temperature_c": temperature_c,
                "temperature_k": temperature_c + 273.15,
                "pressure_mmhg": pressure_mmhg,
                "pressure_kpa": pressure_mmhg * MMHG_TO_KPA,
                "quality_check_1": int(raw["quality_check_1"]),
                "quality_check_2": int(raw["quality_check_2"]),
                "quality_status": _quality_status(raw["quality_check_1"], raw["quality_check_2"]),
                "system_id": " || ".join(ordered_ids),
                "component_order": " || ".join(component_ids),
                "closure_x_abs": abs(sum(x) - 1.0),
                "closure_y_abs": abs(sum(y) - 1.0),
                "smiles_missing_count": sum(row.smiles_status == "missing" for row in component_rows),
                "smiles_parse_failure_count": sum(row.smiles_status == "parse_failed" for row in component_rows),
            }
            for position in range(3):
                record[f"component_id_{position + 1}"] = (
                    component_ids[position] if position < workbook.component_count else ""
                )
                record[f"x{position + 1}"] = x[position] if position < workbook.component_count else np.nan
                record[f"y{position + 1}"] = y[position] if position < workbook.component_count else np.nan
            state_values = [
                record["system_id"],
                round(record["temperature_k"], 9),
                round(record["pressure_kpa"], 9),
                *(round(value, 12) for value in ordered_x),
                *(round(value, 12) for value in ordered_y),
            ]
            record["canonical_state_key"] = repr(tuple(state_values))
            records.append(record)
    frame = pd.DataFrame(records)
    return UnifiedData(workbooks=workbooks, records=frame, occurrences=occurrences)


def ternary_binary_coverage(records: pd.DataFrame) -> pd.DataFrame:
    binary_systems = {
        frozenset(system.split(" || "))
        for system in records.loc[records["dataset"] == "Binary", "system_id"].unique()
    }
    rows = []
    ternary = records.loc[records["dataset"] == "Ternary"].drop_duplicates("system_id")
    for _, row in ternary.iterrows():
        components = tuple(sorted(row["system_id"].split(" || ")))
        pairs = list(combinations(components, 2))
        available = [frozenset(pair) in binary_systems for pair in pairs]
        item: dict[str, Any] = {
            "ternary_system_id": row["system_id"],
            "component_id_1": components[0],
            "component_id_2": components[1],
            "component_id_3": components[2],
            "available_binary_subsystems": int(sum(available)),
            "coverage_category": f"{sum(available)}/3",
        }
        for index, (pair, present) in enumerate(zip(pairs, available), start=1):
            item[f"binary_subsystem_{index}"] = " || ".join(pair)
            item[f"binary_subsystem_{index}_available"] = bool(present)
        rows.append(item)
    return pd.DataFrame(rows).sort_values(
        ["available_binary_subsystems", "ternary_system_id"],
        ascending=[False, True],
    )


def markdown_table(frame: pd.DataFrame, float_format: str = ".4g") -> str:
    if frame.empty:
        return "_No rows._"
    display = frame.copy()
    for column in display.select_dtypes(include=["float"]).columns:
        display[column] = display[column].map(
            lambda value: "" if pd.isna(value) else format(value, float_format)
        )
    headers = [str(column) for column in display.columns]
    lines = [
        "| " + " | ".join(headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    for row in display.itertuples(index=False, name=None):
        values = [str(value).replace("|", "\\|").replace("\n", " ") for value in row]
        lines.append("| " + " | ".join(values) + " |")
    return "\n".join(lines)


def finite(values: pd.Series) -> pd.Series:
    numeric = pd.to_numeric(values, errors="coerce")
    return numeric[np.isfinite(numeric)]


def safe_percentage(count: int, total: int) -> float:
    return 100.0 * count / total if total else math.nan
